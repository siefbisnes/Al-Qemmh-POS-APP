"""
Receipts and the Arabic-language reports are built as real PDF files
using reportlab (see app/services/pdf_utils.py) - NOT a browser engine
anymore. This replaces an earlier Playwright-based implementation:
bundling a full Chromium install into a PyInstaller onefile .exe was
consistently unreliable (path resolution across dev/frozen builds,
version-dependent folder layouts, and PyInstaller import/build
failures). reportlab is pure Python with a single small font asset
bundled alongside it (app/static/fonts/), so there's nothing external
to bundle, resolve a path to, or fail to import.

The receipt/report *_html() functions below are unchanged and still
used for the in-app HTML preview page - only how a downloadable PDF
FILE gets produced changed. The download UX for the user is identical
to before: click, a .pdf file downloads immediately, same filename
convention, same button.
"""
import io

from flask import render_template

from app.services.sales import PAYMENT_LABELS_AR
from app.services.pdf_utils import RTLCanvas, mm_


def _with_labels(payments):
    return [{**p, "label": PAYMENT_LABELS_AR.get(p["method"], p["method"])} for p in payments]


# ---------- store contact info ----------
#
# Same situation as store_name="Al-Qemma" a few lines below: this is
# fixed shop info, not per-transaction data, so it's pinned here rather
# than threaded through every call site. If this ever needs to be
# editable without a code change, move it into app/services/settings.py
# next to warranty_days() etc. (same recommendation as the tax rate
# above) and swap these two lines for a call to it.
STORE_PHONE = "01099170793"
STORE_ADDRESS = (
    "العنوان: المنصورة - شارع بايرو (متفرع من شارع كلية آداب القديمة)، "
    "بعد أول تقاطع بجانب محل مستر هشام للمحمول وأمام محل 200 للملابس - محل القمة."
)


# ---------- receipts ----------

def receipt_html(transaction, auto_print=False):
    txn = dict(transaction)
    txn["payments"] = _with_labels(txn.get("payments", []))
    return render_template(
        "receipts/receipt.html", txn=txn, store_name="Al-Qemma",
        store_phone=STORE_PHONE, store_address=STORE_ADDRESS,
        auto_print=auto_print,
    )


def receipt_pdf_bytes(transaction):
    """Draws the same layout as receipts/receipt.html directly as a PDF
    page: 3-column header, customer box, green-header items table,
    right-aligned totals, payment card, footer. Returns raw PDF bytes,
    same as the old Playwright version did - callers don't need to
    change.

    Handles large orders (many line items, and/or long product names)
    correctly: item names are truncated with an ellipsis to fit their
    column instead of overflowing into neighboring columns, the table
    header is redrawn on every new page, and the totals/payment
    sections check for remaining room and start a fresh page rather
    than silently running off the bottom of the last one."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.colors import HexColor
    from app.services.pdf_utils import truncate_to_fit

    txn = dict(transaction)
    txn["payments"] = _with_labels(txn.get("payments", []))
    invoice_number = txn.get("receipt_number") or f"INV-{(txn.get('created_at') or '----')[:4]}-{txn['id']:06d}"
    lines = txn.get("lines", [])
    subtotal = txn.get("total", 0)
    grand_total = subtotal
    paid = sum(p["amount"] for p in txn["payments"]) if txn.get("payments") else 0
    remaining = grand_total - paid

    buf = io.BytesIO()
    pdf = RTLCanvas(buf, pagesize=A4)
    w, h = pdf.width, pdf.height
    margin = mm_(18)
    right_x = w - margin
    left_x = margin
    top_y = h - mm_(20)
    bottom_margin = mm_(20)

    col_widths = [0.40, 0.13, 0.16, 0.16, 0.15]  # name, qty, price, total, warranty (fractions of table width)
    table_w = right_x - left_x
    col_x = [right_x]
    for frac in col_widths:
        col_x.append(col_x[-1] - table_w * frac)
    headers = ["الصنف", "الكمية", "سعر الوحدة", "الإجمالي", "الضمان"]
    header_h = mm_(7)
    row_h = mm_(7)
    name_col_max_width = table_w * col_widths[0] - mm_(4)  # small padding so text never touches the divider

    def draw_page_header():
        y = top_y
        pdf.draw_center(w / 2, y, "Al-Qemma", font="NotoNaskhArabic-Bold", size=20)
        y -= mm_(7)
        if STORE_PHONE:
            pdf.draw_center(w / 2, y, STORE_PHONE, size=10, color="#5B6472")
        pdf.draw_right(right_x, h - mm_(15), invoice_number, font="NotoNaskhArabic-Bold", size=12, color="#5B6472")
        pdf.draw_right(right_x, h - mm_(21), (txn.get("created_at") or "")[:10], size=10, color="#5B6472")
        pdf.draw_right(right_x, h - mm_(26), (txn.get("created_at") or "")[11:16], size=10, color="#5B6472")
        y = top_y - mm_(16)
        pdf.line(left_x, y, right_x, y)
        return y - mm_(10)

    def draw_table_header(y):
        pdf.c.setFillColor(HexColor("#6F9D8B"))
        pdf.c.rect(col_x[-1], y - header_h, table_w, header_h, stroke=0, fill=1)
        for i, htext in enumerate(headers):
            cx = (col_x[i] + col_x[i + 1]) / 2
            pdf.draw_center(cx, y - mm_(5), htext, font="NotoNaskhArabic-Bold", size=9, color="#FFFFFF")
        return y - header_h

    # ---- header ----
    y = draw_page_header()

    # ---- customer box ----
    box_h = mm_(16)
    pdf.rect(left_x, y - box_h, right_x - left_x, box_h, stroke="#E5E5E5")
    pdf.draw_right(right_x - mm_(4), y - mm_(6), "العميل", size=9, color="#5B6472")
    customer_line = txn.get("customer_name") or "—"
    if txn.get("customer_phone"):
        customer_line += f"   {txn['customer_phone']}"
    pdf.draw_right(right_x - mm_(4), y - mm_(12), customer_line, font="NotoNaskhArabic-Bold", size=11)
    y -= box_h + mm_(8)

    # ---- items table ----
    pdf.draw_right(right_x, y, "الأصناف", font="NotoNaskhArabic-Bold", size=12)
    y -= mm_(6)
    y = draw_table_header(y)

    for line in lines:
        if y - row_h < bottom_margin:
            pdf.new_page()
            y = draw_page_header()
            y = draw_table_header(y)
        name = line.get("service_description") or line.get("product_name") or ""
        name = truncate_to_fit(name, "NotoNaskhArabic", 9, name_col_max_width)
        qty = line.get("quantity", 0)
        price = line.get("selling_price", 0)
        total_line = price * qty
        warranty = f"{line['warranty_days']} يوم" if line.get("warranty_days") else "بدون ضمان"
        values = [name, qty, f"{price:.2f}", f"{total_line:.2f}", warranty]
        for i, val in enumerate(values):
            cx = (col_x[i] + col_x[i + 1]) / 2
            pdf.draw_center(cx, y - mm_(5), val, size=9)
        pdf.line(left_x, y - row_h, right_x, y - row_h, color="#EFEFEF")
        y -= row_h

    y -= mm_(8)

    # ---- totals ----
    totals = [("عدد الأصناف", str(len(lines)), False)]
    totals.append(("الإجمالي الفرعي", f"{subtotal:.2f} ج.م", False))
    totals.append(("الإجمالي", f"{grand_total:.2f} ج.م", True))
    if remaining > 0.009:
        totals.append(("المتبقي", f"{remaining:.2f} ج.م", False))

    totals_block_h = sum((mm_(14) if is_grand else mm_(8)) for _, _, is_grand in totals) + mm_(6)
    if y - totals_block_h < bottom_margin:
        pdf.new_page()
        y = draw_page_header()

    value_col_x = right_x - mm_(50)  # fixed right-aligned column for values, independent of label width
    for label, value, is_grand in totals:
        label_size = 13 if is_grand else 10
        value_size = 18 if is_grand else 10
        row_height = mm_(14) if is_grand else mm_(8)
        if is_grand:
            pdf.line(left_x, y, right_x, y, color="#E5E5E5")
            y -= mm_(6)
        pdf.draw_right(right_x, y, label, size=label_size,
                        color="#1A1F2B" if is_grand else "#5B6472",
                        font="NotoNaskhArabic-Bold" if is_grand else "NotoNaskhArabic")
        pdf.draw_right(value_col_x, y, value, font="NotoNaskhArabic-Bold", size=value_size,
                        color="#6F9D8B" if is_grand else "#1A1F2B")
        y -= row_height

    y -= mm_(6)

    # ---- payment box ----
    if txn.get("payments"):
        pay_h = mm_(8) * len(txn["payments"]) + mm_(4)
        if y - pay_h < bottom_margin:
            pdf.new_page()
            y = draw_page_header()
        pdf.rect(left_x, y - pay_h, right_x - left_x, pay_h, stroke="#E5E5E5", fill="#F7F8F7")
        py = y - mm_(6)
        for p in txn["payments"]:
            pdf.draw_right(right_x - mm_(4), py, p["label"], font="NotoNaskhArabic-Bold", size=10)
            pdf.draw_left(left_x + mm_(4), py, f"{p['amount']:.2f} ج.م", font="NotoNaskhArabic-Bold", size=10)
            py -= mm_(7)
        y -= pay_h + mm_(6)

    # ---- footer ----
    if y < bottom_margin + mm_(20):
        pdf.new_page()
    pdf.draw_center(w / 2, mm_(20), "شكرًا لتعاملكم معنا", font="NotoNaskhArabic-Bold", size=11)
    footer_meta = invoice_number + (f" · {STORE_PHONE}" if STORE_PHONE else "")
    pdf.draw_center(w / 2, mm_(14), footer_meta, size=8, color="#5B6472")

    pdf.save()
    return buf.getvalue()


def receipt_excel_bytes(transaction):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = "إيصال"

    bold = Font(bold=True)
    right = Alignment(horizontal="right")

    ws.append(["Al-Qemma"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["اسم العميل", transaction.get("customer_name") or "—"])
    ws.append(["رقم الهاتف", transaction.get("customer_phone") or "—"])
    ws.append(["التاريخ", (transaction.get("created_at") or "")[:16]])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(["الصنف", "العدد", "السعر"])
    for cell in ws[header_row]:
        cell.font = bold

    for line in transaction.get("lines", []):
        ws.append([line["product_name"], line["quantity"], line["selling_price"]])

    ws.append([])
    ws.append(["الإجمالي", "", transaction.get("total", 0)])
    ws[ws.max_row][0].font = bold

    if transaction.get("payments"):
        ws.append([])
        ws.append(["طريقة الدفع", "المبلغ"])
        for cell in ws[ws.max_row]:
            cell.font = bold
        for p in _with_labels(transaction["payments"]):
            ws.append([p["label"], p["amount"]])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = right
    for col, width in zip("ABC", (28, 14, 14)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------- Arabic reports ----------

def report_html(report, date_from=None, date_to=None, auto_print=False):
    return render_template(
        "receipts/report.html",
        report=report, date_from=date_from, date_to=date_to, auto_print=auto_print,
    )


def report_pdf_bytes(report, date_from=None, date_to=None):
    """Same cash/online/total/expenses summary as report_excel_bytes(),
    drawn as a simple labeled PDF card instead of an Excel sheet."""
    return _simple_kv_report_pdf(
        title="ملخص التقرير المالي",
        subtitle=f"{date_from or 'البداية'} — {date_to or 'الآن'}" if (date_from or date_to) else None,
        rows=[
            ("نقدي", f"{report['cash']:.2f} ج.م"),
            ("أونلاين", f"{report['online']:.2f} ج.م"),
            ("الإجمالي", f"{report['total']:.2f} ج.م"),
            ("المصروفات", f"{report['expenses']:.2f} ج.م"),
        ],
    )


def report_excel_bytes(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    bold = Font(bold=True)
    right = Alignment(horizontal="right")

    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True
    ws.title = "ملخص"
    ws.append(["المؤشر", "القيمة"])
    for cell in ws[1]:
        cell.font = bold
    for label, value in [
        ("نقدي", report["cash"]),
        ("أونلاين", report["online"]),
        ("الإجمالي", report["total"]),
        ("المصروفات", report["expenses"]),
    ]:
        ws.append([label, value])

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = right
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------- تقرير اليوم (today report) ----------

def today_report_html(report, auto_print=False):
    return render_template("receipts/today_report.html", report=report, store_name="Al-Qemma", auto_print=auto_print)


def today_report_pdf_bytes(report):
    """report here is the dict from report_service.today_report():
    {"drawer": {...}, "today_total": ...} per app/routes/reports.py.
    Flattened into simple label/value rows - if today_report.html (not
    available to verify field-by-field against) shows additional
    figures beyond these, add rows here to match; the structure below
    covers every field this module has seen referenced elsewhere in the
    app (drawer.*, today_total)."""
    drawer = report.get("drawer") or {}
    rows = []
    if isinstance(drawer, dict):
        for key, value in drawer.items():
            if isinstance(value, (int, float)):
                rows.append((key, f"{value:.2f} ج.م"))
    if "today_total" in report:
        rows.append(("اموال اليوم", f"{report['today_total']:.2f} ج.م"))
    return _simple_kv_report_pdf(title="تقرير اليوم", subtitle=None, rows=rows)


def _simple_kv_report_pdf(title, subtitle, rows):
    """Shared layout for the two summary-card style reports above:
    centered title, optional subtitle, a right-aligned label / left-
    aligned value row per entry with light divider lines."""
    from reportlab.lib.pagesizes import A4

    buf = io.BytesIO()
    pdf = RTLCanvas(buf, pagesize=A4)
    w, h = pdf.width, pdf.height
    margin = mm_(20)
    right_x = w - margin
    left_x = margin
    y = h - mm_(30)

    pdf.draw_center(w / 2, y, "Al-Qemma", font="NotoNaskhArabic-Bold", size=18)
    y -= mm_(10)
    pdf.draw_center(w / 2, y, title, font="NotoNaskhArabic-Bold", size=14, color="#6F9D8B")
    y -= mm_(8)
    if subtitle:
        pdf.draw_center(w / 2, y, subtitle, size=10, color="#5B6472")
        y -= mm_(8)
    pdf.line(left_x, y, right_x, y)
    y -= mm_(12)

    for label, value in rows:
        pdf.draw_right(right_x, y, label, font="NotoNaskhArabic-Bold", size=12)
        pdf.draw_left(left_x, y, value, size=12, color="#1A1F2B")
        y -= mm_(6)
        pdf.line(left_x, y, right_x, y, color="#EFEFEF")
        y -= mm_(6)

    pdf.save()
    return buf.getvalue()