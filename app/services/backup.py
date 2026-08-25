"""
A full export of everything in the database, for three purposes the brief
calls out specifically: backing up before an update, moving the system to
another computer, and disaster recovery. Three formats, same underlying
data:

- Raw .db: an exact, consistent copy of the SQLite file itself. Using
  sqlite3's own backup() API instead of just copying the file means this is
  safe even if (in theory) something else has a transaction open at the
  moment of export - SQLite handles that correctly, a plain file copy
  would not.
- Excel: every table as its own sheet, human-readable, easy to open and
  skim without needing SQLite tooling.
- CSV: one .csv per table, zipped together - the most portable option,
  importable into literally anything.
"""
import csv
import io
import sqlite3
import zipfile

from flask import current_app


def backup_db_bytes():
    """A consistent, exact copy of the live database file, taken via
    SQLite's own backup API (safe even if something else has a
    transaction open at that exact moment) rather than a raw file copy."""
    import os
    import tempfile

    fd, tmp_path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    os.remove(tmp_path)  # backup() needs to create the destination itself

    src = sqlite3.connect(current_app.config["DATABASE_PATH"])
    dst = sqlite3.connect(tmp_path)
    try:
        src.backup(dst)
    finally:
        dst.close()
        src.close()

    with open(tmp_path, "rb") as f:
        data = f.read()
    os.remove(tmp_path)

    return io.BytesIO(data), "alqemma_backup.db"


def _all_tables():
    return [
        "categories", "category_fields", "products", "specifications",
        "product_images", "compatibility", "customers",
        "transactions", "sales", "sale_payments", "warranties",
        "settings", "expenses", "purchases", "manual_adjustments",
        "stock_writeoffs",
    ]


def _fetch_table(cur, table):
    cur.execute(f"SELECT * FROM {table}")
    columns = [d[0] for d in cur.description]
    rows = [dict(zip(columns, row)) for row in cur.fetchall()]
    return columns, rows


def export_excel_bytes():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    conn = sqlite3.connect(current_app.config["DATABASE_PATH"])
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    wb = Workbook()
    wb.remove(wb.active)

    for table in _all_tables():
        columns, rows = _fetch_table(cur, table)
        ws = wb.create_sheet(table[:31])  # Excel sheet name limit
        ws.append(columns)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row[c] for c in columns])

    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "alqemma_full_export.xlsx"


def export_csv_zip_bytes():
    conn = sqlite3.connect(current_app.config["DATABASE_PATH"])
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for table in _all_tables():
            columns, rows = _fetch_table(cur, table)
            csv_buf = io.StringIO()
            writer = csv.writer(csv_buf)
            writer.writerow(columns)
            for row in rows:
                writer.writerow([row[c] for c in columns])
            zf.writestr(f"{table}.csv", csv_buf.getvalue())

    conn.close()
    zip_buf.seek(0)
    return zip_buf, "alqemma_full_export_csv.zip"

